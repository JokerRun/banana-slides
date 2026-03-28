import importlib.util
import json
import sqlite3
import sys
from pathlib import Path

from openpyxl import load_workbook


SCRIPT_PATH = Path(__file__).resolve().parents[3] / 'data' / 'debug' / 'export_restyle_debug_excel.py'


def load_exporter_module():
    spec = importlib.util.spec_from_file_location('export_restyle_debug_excel', SCRIPT_PATH)
    module = importlib.util.module_from_spec(spec)
    assert spec.loader is not None
    sys.modules[spec.name] = module
    spec.loader.exec_module(module)
    return module


def write_json(path: Path, payload: dict):
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding='utf-8')


def build_sqlite_db(db_path: Path):
    conn = sqlite3.connect(db_path)
    try:
        conn.executescript(
            '''
            CREATE TABLE projects (
                id TEXT PRIMARY KEY,
                idea_prompt TEXT,
                creation_type TEXT,
                status TEXT,
                restyle_prompt TEXT,
                created_at TEXT,
                updated_at TEXT
            );
            CREATE TABLE pages (
                id TEXT PRIMARY KEY,
                project_id TEXT,
                order_index INTEGER,
                status TEXT,
                generated_image_path TEXT,
                cached_image_path TEXT,
                original_slide_image_path TEXT,
                restyle_base_prompt_snapshot TEXT
            );
            CREATE TABLE tasks (
                id TEXT PRIMARY KEY,
                project_id TEXT,
                task_type TEXT,
                status TEXT,
                progress TEXT,
                error_message TEXT,
                created_at TEXT,
                completed_at TEXT
            );
            CREATE TABLE page_image_versions (
                id TEXT PRIMARY KEY,
                page_id TEXT,
                image_path TEXT,
                version_number INTEGER,
                is_current INTEGER,
                created_at TEXT
            );
            '''
        )
        conn.execute(
            'INSERT INTO projects VALUES (?, ?, ?, ?, ?, ?, ?)',
            ('project-1', 'Demo Project', 'restyle', 'COMPLETED', 'Use dark theme', '2026-03-28T10:00:00', '2026-03-28T10:30:00'),
        )
        conn.execute(
            'INSERT INTO pages VALUES (?, ?, ?, ?, ?, ?, ?, ?)',
            (
                'page-1',
                'project-1',
                0,
                'COMPLETED',
                'project-1/pages/page-1_v2.png',
                'project-1/cache/page-1_v2.jpg',
                'project-1/pages/originals/page-1.png',
                'BASE SNAPSHOT',
            ),
        )
        conn.executemany(
            'INSERT INTO tasks VALUES (?, ?, ?, ?, ?, ?, ?, ?)',
            [
                ('task-batch', 'project-1', 'RESTYLE_IMAGES', 'COMPLETED', json.dumps({'completed': 1, 'failed': 0}), None, '2026-03-28T10:05:00', '2026-03-28T10:15:00'),
                ('task-edit', 'project-1', 'EDIT_PAGE_IMAGE', 'COMPLETED', json.dumps({'completed': 1, 'failed': 0}), None, '2026-03-28T10:20:00', '2026-03-28T10:25:00'),
            ],
        )
        conn.executemany(
            'INSERT INTO page_image_versions VALUES (?, ?, ?, ?, ?, ?)',
            [
                ('v1', 'page-1', 'project-1/pages/page-1_v1.png', 1, 0, '2026-03-28T10:15:00'),
                ('v2', 'page-1', 'project-1/pages/page-1_v2.png', 2, 1, '2026-03-28T10:25:00'),
            ],
        )
        conn.commit()
    finally:
        conn.close()


def build_debug_tree(debug_root: Path):
    write_json(
        debug_root / 'task-batch' / 'task' / 'summary.json',
        {
            'trace': {'task_id': 'task-batch', 'project_id': 'project-1', 'flow_kind': 'first_pass_restyle'},
            'event_name': 'summary',
            'event': {
                'status': 'COMPLETED',
                'completed': 1,
                'failed': 0,
                'total_pages': 1,
                'page_results': [
                    {
                        'page_id': 'page-1',
                        'page_order_index': 1,
                        'version_number': 1,
                        'image_path': 'project-1/pages/page-1_v1.png',
                        'error': None,
                    }
                ],
            },
        },
    )
    write_json(
        debug_root / 'task-batch' / 'task' / 'started.json',
        {
            'trace': {'task_id': 'task-batch', 'project_id': 'project-1', 'flow_kind': 'first_pass_restyle'},
            'event_name': 'started',
            'event': {'total_pages': 1, 'style_ref_count': 1},
        },
    )
    page_dir = debug_root / 'task-batch' / 'pages' / 'page-001-page-1'
    write_json(
        page_dir / 'context_built.json',
        {
            'trace': {'task_id': 'task-batch', 'project_id': 'project-1', 'page_id': 'page-1', 'page_order_index': 1, 'flow_kind': 'first_pass_restyle'},
            'event_name': 'context_built',
            'event': {
                'prompt_len': 123,
                'snapshot_present': False,
                'image_manifest': [{'kind': 'original_slide', 'path': '/app/uploads/project-1/pages/originals/page-1.png'}],
            },
        },
    )
    write_json(
        page_dir / 'provider_decision.json',
        {
            'trace': {'task_id': 'task-batch', 'project_id': 'project-1', 'page_id': 'page-1', 'page_order_index': 1, 'flow_kind': 'first_pass_restyle'},
            'event_name': 'provider_decision',
            'event': {'provider': 'GenAIProvider', 'model': 'gemini-image', 'thinking_level': 'none'},
        },
    )
    write_json(
        page_dir / 'provider_request.json',
        {
            'trace': {
                'task_id': 'task-batch',
                'project_id': 'project-1',
                'page_id': 'page-1',
                'page_order_index': 1,
                'page_version_number': None,
                'source_version_number': None,
                'flow_kind': 'first_pass_restyle',
            },
            'event_name': 'provider_request',
            'event': {
                'provider': 'GenAIProvider',
                'model': 'gemini-image',
                'thinking_level': 'none',
                'prompt': 'FULL FIRST PASS PROMPT',
                'prompt_len': 22,
                'aspect_ratio': '16:9',
                'resolution': '2K',
                'ref_image_paths': ['/app/uploads/project-1/pages/originals/page-1.png'],
            },
        },
    )
    write_json(
        page_dir / 'provider_result.json',
        {
            'trace': {'task_id': 'task-batch', 'project_id': 'project-1', 'page_id': 'page-1', 'page_order_index': 1, 'flow_kind': 'first_pass_restyle'},
            'event_name': 'provider_result',
            'event': {'elapsed_seconds': 12.5, 'result_image_size': [1920, 1080], 'error_stage': None},
        },
    )
    write_json(
        page_dir / 'saved_version.json',
        {
            'trace': {
                'task_id': 'task-batch',
                'project_id': 'project-1',
                'page_id': 'page-1',
                'page_order_index': 1,
                'page_version_number': 1,
                'source_version_number': None,
                'flow_kind': 'first_pass_restyle',
            },
            'event_name': 'saved_version',
            'event': {'image_path': 'project-1/pages/page-1_v1.png', 'version_number': 1, 'snapshot_persisted': True},
        },
    )
    edit_dir = debug_root / 'task-edit'
    write_json(
        edit_dir / 'context_built.json',
        {
            'trace': {
                'task_id': 'task-edit',
                'project_id': 'project-1',
                'page_id': 'page-1',
                'page_order_index': 1,
                'page_version_number': None,
                'source_version_number': 1,
                'flow_kind': 'edit_restyle',
            },
            'event_name': 'context_built',
            'event': {
                'snapshot_source': 'persisted',
                'degraded_context': False,
                'turns_summary': [{'role': 'user', 'text_len': 10}],
                'image_manifest': [{'kind': 'current_version', 'path': '/app/uploads/project-1/pages/page-1_v1.png'}],
            },
        },
    )
    write_json(
        edit_dir / 'provider_decision.json',
        {
            'trace': {'task_id': 'task-edit', 'project_id': 'project-1', 'page_id': 'page-1', 'page_order_index': 1, 'source_version_number': 1, 'flow_kind': 'edit_restyle'},
            'event_name': 'provider_decision',
            'event': {'provider': 'GenAIProvider', 'model': 'gemini-image', 'thinking_level': 'low'},
        },
    )
    write_json(
        edit_dir / 'provider_request.json',
        {
            'trace': {
                'task_id': 'task-edit',
                'project_id': 'project-1',
                'page_id': 'page-1',
                'page_order_index': 1,
                'page_version_number': None,
                'source_version_number': 1,
                'flow_kind': 'edit_restyle',
            },
            'event_name': 'provider_request',
            'event': {
                'context_mode': 'conversation',
                'snapshot_source': 'persisted',
                'turns_summary': [{'role': 'user', 'text_len': 12}],
                'image_manifest': [{'kind': 'current_version', 'path': '/app/uploads/project-1/pages/page-1_v1.png'}],
                'conversation_contents': [{'role': 'user', 'parts': [{'text': 'FULL EDIT REQUEST'}]}],
            },
        },
    )
    write_json(
        edit_dir / 'provider_result.json',
        {
            'trace': {'task_id': 'task-edit', 'project_id': 'project-1', 'page_id': 'page-1', 'page_order_index': 1, 'source_version_number': 1, 'flow_kind': 'edit_restyle'},
            'event_name': 'provider_result',
            'event': {'elapsed_seconds': 6.25, 'error_stage': None},
        },
    )
    write_json(
        edit_dir / 'saved_version.json',
        {
            'trace': {
                'task_id': 'task-edit',
                'project_id': 'project-1',
                'page_id': 'page-1',
                'page_order_index': 1,
                'page_version_number': None,
                'source_version_number': 1,
                'flow_kind': 'edit_restyle',
            },
            'event_name': 'saved_version',
            'event': {'image_path': 'project-1/pages/page-1_v2.png', 'version_number': 2},
        },
    )


def test_export_restyle_debug_excel_creates_overview_and_flat_workbooks(tmp_path):
    db_path = tmp_path / 'database.db'
    debug_root = tmp_path / 'restyle-context'
    uploads_root = tmp_path / 'uploads'
    output_dir = tmp_path / 'export'
    uploads_root.mkdir()
    (uploads_root / 'project-1' / 'pages' / 'originals').mkdir(parents=True)
    (uploads_root / 'project-1' / 'pages' / 'originals' / 'page-1.png').write_bytes(b'png')
    (uploads_root / 'project-1' / 'pages' / 'page-1_v1.png').parent.mkdir(parents=True, exist_ok=True)
    (uploads_root / 'project-1' / 'pages' / 'page-1_v1.png').write_bytes(b'v1')
    (uploads_root / 'project-1' / 'pages' / 'page-1_v2.png').write_bytes(b'v2')
    (uploads_root / 'project-1' / 'cache').mkdir(parents=True)
    (uploads_root / 'project-1' / 'cache' / 'page-1_v2.jpg').write_bytes(b'jpg')

    build_sqlite_db(db_path)
    build_debug_tree(debug_root)

    exporter = load_exporter_module()
    outputs = exporter.export_debug_workbooks(
        debug_root=debug_root,
        db_path=db_path,
        uploads_root=uploads_root,
        output_dir=output_dir,
    )

    overview_path = Path(outputs['overview'])
    flat_path = Path(outputs['flat'])
    assert overview_path.exists()
    assert flat_path.exists()

    overview_wb = load_workbook(overview_path)
    assert overview_wb.sheetnames == ['Projects', 'Tasks', 'Pages', 'Versions', 'Requests']

    projects_sheet = overview_wb['Projects']
    assert projects_sheet['A2'].value == 'project-1'
    assert projects_sheet['B2'].value == 'Demo Project'

    versions_sheet = overview_wb['Versions']
    headers = [cell.value for cell in versions_sheet[1]]
    assert 'version_number' in headers
    assert 'prompt_full' in headers
    rows = list(versions_sheet.iter_rows(min_row=2, values_only=True))
    assert len(rows) == 2
    version_numbers = {row[headers.index('version_number')] for row in rows}
    assert version_numbers == {1, 2}
    prompt_values = {row[headers.index('prompt_full')] for row in rows}
    assert 'FULL FIRST PASS PROMPT' in prompt_values
    request_modes = {row[headers.index('request_kind')] for row in rows}
    assert request_modes == {'first_pass', 'edit'}

    requests_sheet = overview_wb['Requests']
    request_headers = [cell.value for cell in requests_sheet[1]]
    request_rows = list(requests_sheet.iter_rows(min_row=2, values_only=True))
    assert any(row[request_headers.index('conversation_json_full')] for row in request_rows)

    flat_wb = load_workbook(flat_path)
    assert flat_wb.sheetnames == ['FlatExport']
    flat_sheet = flat_wb['FlatExport']
    flat_headers = [cell.value for cell in flat_sheet[1]]
    assert 'project_name' in flat_headers
    assert 'request_json_full' in flat_headers
    flat_rows = list(flat_sheet.iter_rows(min_row=2, values_only=True))
    assert len(flat_rows) == 2
