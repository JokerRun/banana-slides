from __future__ import annotations

import argparse
import json
import sqlite3
from collections import defaultdict
from dataclasses import dataclass
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


HEADER_FILL = PatternFill('solid', fgColor='D9EAF7')
HEADER_FONT = Font(name='Arial', size=10, bold=True)
BODY_FONT = Font(name='Arial', size=10)
WRAP_ALIGNMENT = Alignment(vertical='top', wrap_text=True)
TOP_ALIGNMENT = Alignment(vertical='top')

OVERVIEW_SHEETS = ['Projects', 'Tasks', 'Pages', 'Versions', 'Requests']
LONG_TEXT_COLUMNS = {
    'restyle_prompt',
    'restyle_base_prompt_snapshot',
    'prompt_full',
    'prompt_preview',
    'request_json_full',
    'request_event_json_full',
    'provider_result_json_full',
    'saved_version_json_full',
    'conversation_json_full',
    'conversation_preview',
    'turns_summary_json_full',
    'image_manifest_json_full',
    'ref_image_paths_json',
    'error_message',
}

PROJECT_HEADERS = [
    'project_id',
    'project_name',
    'creation_type',
    'project_status',
    'restyle_prompt',
    'total_pages',
    'task_count',
    'version_count',
    'current_version_count',
    'has_snapshot_pages',
    'first_task_created_at',
    'last_task_created_at',
]
TASK_HEADERS = [
    'task_id',
    'project_id',
    'project_name',
    'task_type',
    'task_status',
    'flow_kind',
    'request_kind',
    'page_count',
    'completed',
    'failed',
    'error_message',
    'created_at',
    'completed_at',
    'debug_path',
    'has_started_json',
    'has_summary_json',
    'has_saved_version',
]
PAGE_HEADERS = [
    'project_id',
    'project_name',
    'page_id',
    'order_index',
    'page_status',
    'generated_image_path',
    'generated_image_host_path',
    'generated_image_exists',
    'cached_image_path',
    'cached_image_host_path',
    'cached_image_exists',
    'original_slide_image_path',
    'original_slide_host_path',
    'original_slide_exists',
    'restyle_base_prompt_snapshot',
    'snapshot_len',
    'current_version_number',
    'current_version_image_path',
    'current_version_host_path',
    'current_version_exists',
    'version_count',
]
VERSION_HEADERS = [
    'project_id',
    'project_name',
    'project_status',
    'page_id',
    'order_index',
    'page_status',
    'version_number',
    'is_current',
    'image_path',
    'host_image_path',
    'image_exists',
    'source_task_id',
    'source_task_type',
    'task_status',
    'request_kind',
    'flow_kind',
    'source_version_number',
    'snapshot_persisted',
    'snapshot_source',
    'provider',
    'model',
    'thinking_level',
    'elapsed_seconds',
    'result_image_size',
    'error_stage',
    'error_message',
    'prompt_full',
    'prompt_preview',
    'request_json_full',
    'conversation_json_full',
    'image_manifest_json_full',
    'turns_summary_json_full',
]
REQUEST_HEADERS = [
    'task_id',
    'project_id',
    'project_name',
    'page_id',
    'order_index',
    'version_number',
    'request_kind',
    'flow_kind',
    'provider',
    'model',
    'thinking_level',
    'prompt_len',
    'prompt_full',
    'prompt_preview',
    'request_json_full',
    'request_event_json_full',
    'conversation_json_full',
    'conversation_preview',
    'ref_image_paths_json',
    'image_manifest_json_full',
    'turns_summary_json_full',
    'snapshot_source',
    'degraded_context',
    'provider_fallback',
    'elapsed_seconds',
    'error_stage',
    'error_message',
    'provider_result_json_full',
    'saved_version_json_full',
]
FLAT_HEADERS = [
    'project_id',
    'project_name',
    'creation_type',
    'project_status',
    'restyle_prompt',
    'task_id',
    'task_type',
    'task_status',
    'flow_kind',
    'request_kind',
    'page_id',
    'order_index',
    'page_status',
    'version_number',
    'is_current',
    'source_version_number',
    'image_path',
    'host_image_path',
    'image_exists',
    'provider',
    'model',
    'thinking_level',
    'prompt_len',
    'prompt_preview',
    'prompt_full',
    'snapshot_source',
    'snapshot_persisted',
    'elapsed_seconds',
    'result_image_size',
    'error_stage',
    'error_message',
    'request_json_full',
    'conversation_json_full',
    'image_manifest_json_full',
    'turns_summary_json_full',
]


@dataclass
class DebugEntry:
    task_id: str
    project_id: str | None
    page_id: str | None
    order_index: int | None
    flow_kind: str | None
    request_kind: str
    version_number: int | None
    source_version_number: int | None
    image_path: str | None
    prompt_len: int | None
    prompt_full: str | None
    provider: str | None
    model: str | None
    thinking_level: str | None
    ref_image_paths_json: str | None
    image_manifest_json_full: str | None
    turns_summary_json_full: str | None
    conversation_json_full: str | None
    snapshot_source: str | None
    degraded_context: bool | None
    provider_fallback: bool | None
    elapsed_seconds: float | None
    result_image_size: str | None
    error_stage: str | None
    error_message: str | None
    snapshot_persisted: bool | None
    debug_path: str
    request_json_full: str | None
    request_event_json_full: str | None
    provider_result_json_full: str | None
    saved_version_json_full: str | None


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description='Export restyle debug data to Excel workbooks.')
    parser.add_argument('--debug-root', default='data/debug/restyle-context')
    parser.add_argument('--db', default='data/instance/database.db')
    parser.add_argument('--uploads-root', default='data/uploads')
    parser.add_argument('--output-dir', default='data/debug/export')
    parser.add_argument('--project-id', action='append', dest='project_ids')
    parser.add_argument('--task-id', action='append', dest='task_ids')
    parser.add_argument('--latest', type=int, default=None, help='Keep only latest N tasks after filtering')
    return parser.parse_args()


def load_json(path: Path) -> dict[str, Any]:
    return json.loads(path.read_text(encoding='utf-8'))


def json_text(value: Any) -> str | None:
    if value is None:
        return None
    return json.dumps(value, ensure_ascii=False, indent=2, sort_keys=True)


def safe_int(value: Any) -> int | None:
    if value is None or value == '':
        return None
    try:
        return int(value)
    except (TypeError, ValueError):
        return None


def prompt_preview(value: str | None, limit: int = 160) -> str | None:
    if not value:
        return value
    compact = ' '.join(value.split())
    return compact if len(compact) <= limit else compact[: limit - 3] + '...'


def extract_prompt(event: dict[str, Any]) -> str | None:
    prompt = event.get('prompt')
    if isinstance(prompt, str) and prompt.strip():
        return prompt
    conversation = event.get('conversation_contents') or []
    texts: list[str] = []
    for turn in conversation:
        for part in turn.get('parts', []):
            text = part.get('text')
            if isinstance(text, str) and text.strip():
                texts.append(text)
    if not texts:
        return None
    return '\n\n---\n\n'.join(texts)


def host_path(path_value: str | None, uploads_root: Path) -> Path | None:
    if not path_value:
        return None
    if path_value.startswith('/app/uploads/'):
        return uploads_root / path_value.removeprefix('/app/uploads/')
    path = Path(path_value)
    if path.is_absolute():
        return path
    return uploads_root / path_value


def connect_readonly(db_path: Path) -> sqlite3.Connection:
    uri = f'file:{db_path.resolve()}?mode=ro'
    conn = sqlite3.connect(uri, uri=True)
    conn.row_factory = sqlite3.Row
    return conn


def fetch_rows(conn: sqlite3.Connection, query: str, params: tuple[Any, ...] = ()) -> list[dict[str, Any]]:
    return [dict(row) for row in conn.execute(query, params).fetchall()]


def load_database(db_path: Path) -> dict[str, Any]:
    conn = connect_readonly(db_path)
    try:
        projects = fetch_rows(
            conn,
            '''
            SELECT id, idea_prompt, creation_type, status, restyle_prompt, created_at, updated_at
            FROM projects
            '''
        )
        pages = fetch_rows(
            conn,
            '''
            SELECT id, project_id, order_index, status, generated_image_path, cached_image_path,
                   original_slide_image_path, restyle_base_prompt_snapshot
            FROM pages
            '''
        )
        tasks = fetch_rows(
            conn,
            '''
            SELECT id, project_id, task_type, status, progress, error_message, created_at, completed_at
            FROM tasks
            '''
        )
        versions = fetch_rows(
            conn,
            '''
            SELECT id, page_id, image_path, version_number, is_current, created_at
            FROM page_image_versions
            '''
        )
    finally:
        conn.close()

    projects_by_id = {row['id']: row for row in projects}
    pages_by_id = {row['id']: row for row in pages}
    tasks_by_id = {row['id']: row for row in tasks}
    versions_by_page: dict[str, list[dict[str, Any]]] = defaultdict(list)
    current_versions: dict[str, dict[str, Any]] = {}
    for row in versions:
        versions_by_page[row['page_id']].append(row)
        if row['is_current']:
            current_versions[row['page_id']] = row
    for rows in versions_by_page.values():
        rows.sort(key=lambda item: safe_int(item.get('version_number')) or 0)

    pages_by_project: dict[str, list[dict[str, Any]]] = defaultdict(list)
    tasks_by_project: dict[str, list[dict[str, Any]]] = defaultdict(list)
    for row in pages:
        pages_by_project[row['project_id']].append(row)
    for row in tasks:
        tasks_by_project[row['project_id']].append(row)
    for rows in pages_by_project.values():
        rows.sort(key=lambda item: safe_int(item.get('order_index')) or 0)
    for rows in tasks_by_project.values():
        rows.sort(key=lambda item: item.get('created_at') or '')

    return {
        'projects': projects,
        'projects_by_id': projects_by_id,
        'pages': pages,
        'pages_by_id': pages_by_id,
        'pages_by_project': pages_by_project,
        'tasks': tasks,
        'tasks_by_id': tasks_by_id,
        'tasks_by_project': tasks_by_project,
        'versions': versions,
        'versions_by_page': versions_by_page,
        'current_versions': current_versions,
    }


def detect_task_kind(task_dir: Path) -> str | None:
    if (task_dir / 'task').exists() or (task_dir / 'pages').exists():
        return 'batch'
    if (task_dir / 'provider_request.json').exists() or (task_dir / 'saved_version.json').exists():
        return 'edit'
    return None


def build_entry(task_id: str, artifact_dir: Path, request_kind: str) -> DebugEntry | None:
    request_path = artifact_dir / 'provider_request.json'
    result_path = artifact_dir / 'provider_result.json'
    saved_version_path = artifact_dir / 'saved_version.json'
    context_path = artifact_dir / 'context_built.json'
    decision_path = artifact_dir / 'provider_decision.json'

    request_payload = load_json(request_path) if request_path.exists() else None
    result_payload = load_json(result_path) if result_path.exists() else None
    saved_payload = load_json(saved_version_path) if saved_version_path.exists() else None
    context_payload = load_json(context_path) if context_path.exists() else None
    decision_payload = load_json(decision_path) if decision_path.exists() else None

    traces = [payload.get('trace', {}) for payload in [request_payload, result_payload, saved_payload, context_payload, decision_payload] if payload]
    if not traces:
        return None
    trace = next((candidate for candidate in traces if candidate), {})
    request_event = (request_payload or {}).get('event', {})
    result_event = (result_payload or {}).get('event', {})
    saved_event = (saved_payload or {}).get('event', {})
    context_event = (context_payload or {}).get('event', {})
    decision_event = (decision_payload or {}).get('event', {})

    version_number = safe_int(saved_event.get('version_number'))
    result_image_size = result_event.get('result_image_size')

    return DebugEntry(
        task_id=task_id,
        project_id=trace.get('project_id'),
        page_id=trace.get('page_id'),
        order_index=safe_int(trace.get('page_order_index')),
        flow_kind=trace.get('flow_kind'),
        request_kind=request_kind,
        version_number=version_number,
        source_version_number=safe_int(trace.get('source_version_number')),
        image_path=saved_event.get('image_path'),
        prompt_len=safe_int(request_event.get('prompt_len')),
        prompt_full=extract_prompt(request_event),
        provider=request_event.get('provider') or decision_event.get('provider'),
        model=request_event.get('model') or decision_event.get('model'),
        thinking_level=request_event.get('thinking_level') or decision_event.get('thinking_level'),
        ref_image_paths_json=json_text(request_event.get('ref_image_paths')),
        image_manifest_json_full=json_text(request_event.get('image_manifest') or context_event.get('image_manifest')),
        turns_summary_json_full=json_text(request_event.get('turns_summary') or context_event.get('turns_summary')),
        conversation_json_full=json_text(request_event.get('conversation_contents')),
        snapshot_source=request_event.get('snapshot_source') or context_event.get('snapshot_source'),
        degraded_context=context_event.get('degraded_context') if context_payload else request_event.get('degraded_context'),
        provider_fallback=result_event.get('provider_fallback'),
        elapsed_seconds=result_event.get('elapsed_seconds'),
        result_image_size=json_text(result_image_size),
        error_stage=result_event.get('error_stage'),
        error_message=result_event.get('error_message'),
        snapshot_persisted=saved_event.get('snapshot_persisted'),
        debug_path=str(artifact_dir),
        request_json_full=json_text(request_payload),
        request_event_json_full=json_text(request_event),
        provider_result_json_full=json_text(result_payload),
        saved_version_json_full=json_text(saved_payload),
    )


def load_debug_entries(debug_root: Path) -> tuple[list[DebugEntry], dict[str, dict[str, Any]]]:
    entries: list[DebugEntry] = []
    task_meta: dict[str, dict[str, Any]] = {}

    for task_dir in sorted(path for path in debug_root.iterdir() if path.is_dir()):
        task_id = task_dir.name
        task_kind = detect_task_kind(task_dir)
        if not task_kind:
            continue

        meta = {
            'task_id': task_id,
            'debug_path': str(task_dir),
            'task_kind': task_kind,
            'has_started_json': False,
            'has_summary_json': False,
            'has_saved_version': False,
            'flow_kind': None,
            'request_kind': 'first_pass' if task_kind == 'batch' else 'edit',
            'page_count': None,
            'completed': None,
            'failed': None,
            'project_id': None,
        }

        if task_kind == 'batch':
            started_path = task_dir / 'task' / 'started.json'
            summary_path = task_dir / 'task' / 'summary.json'
            if started_path.exists():
                started_payload = load_json(started_path)
                meta['has_started_json'] = True
                meta['flow_kind'] = started_payload.get('trace', {}).get('flow_kind')
                meta['project_id'] = started_payload.get('trace', {}).get('project_id')
                meta['page_count'] = started_payload.get('event', {}).get('total_pages')
            if summary_path.exists():
                summary_payload = load_json(summary_path)
                meta['has_summary_json'] = True
                meta['flow_kind'] = meta['flow_kind'] or summary_payload.get('trace', {}).get('flow_kind')
                meta['project_id'] = meta['project_id'] or summary_payload.get('trace', {}).get('project_id')
                meta['completed'] = summary_payload.get('event', {}).get('completed')
                meta['failed'] = summary_payload.get('event', {}).get('failed')
                meta['page_count'] = meta['page_count'] or summary_payload.get('event', {}).get('total_pages')

            pages_root = task_dir / 'pages'
            for artifact_dir in sorted(path for path in pages_root.iterdir() if path.is_dir()):
                entry = build_entry(task_id, artifact_dir, 'first_pass')
                if entry:
                    entries.append(entry)
                    meta['has_saved_version'] = meta['has_saved_version'] or bool(entry.version_number)
                    meta['project_id'] = meta['project_id'] or entry.project_id
                    meta['flow_kind'] = meta['flow_kind'] or entry.flow_kind
        else:
            entry = build_entry(task_id, task_dir, 'edit')
            if entry:
                entries.append(entry)
                meta['has_saved_version'] = bool(entry.version_number)
                meta['project_id'] = entry.project_id
                meta['flow_kind'] = entry.flow_kind

        task_meta[task_id] = meta

    return entries, task_meta


def apply_filters(
    db_data: dict[str, Any],
    debug_entries: list[DebugEntry],
    task_meta: dict[str, dict[str, Any]],
    project_ids: set[str] | None,
    task_ids: set[str] | None,
    latest: int | None,
) -> tuple[list[DebugEntry], dict[str, dict[str, Any]], dict[str, Any]]:
    allowed_tasks: set[str] = set(task_meta)
    if task_ids:
        allowed_tasks &= task_ids

    if project_ids:
        allowed_tasks &= {
            task_id for task_id, meta in task_meta.items() if meta.get('project_id') in project_ids
        } | {
            row['id'] for row in db_data['tasks'] if row['project_id'] in project_ids
        }

    if latest:
        sortable = []
        for task_id in allowed_tasks:
            task_row = db_data['tasks_by_id'].get(task_id)
            created_at = (task_row or {}).get('created_at') or ''
            sortable.append((created_at, task_id))
        sortable.sort()
        allowed_tasks = {task_id for _, task_id in sortable[-latest:]}

    filtered_entries = [entry for entry in debug_entries if entry.task_id in allowed_tasks]
    filtered_meta = {task_id: meta for task_id, meta in task_meta.items() if task_id in allowed_tasks}

    if not filtered_entries and not filtered_meta and not project_ids and not task_ids:
        return debug_entries, task_meta, db_data

    relevant_project_ids = set(project_ids or [])
    for entry in filtered_entries:
        if entry.project_id:
            relevant_project_ids.add(entry.project_id)
    for task_id in filtered_meta:
        task_row = db_data['tasks_by_id'].get(task_id)
        if task_row and task_row['project_id']:
            relevant_project_ids.add(task_row['project_id'])

    projects = [row for row in db_data['projects'] if not relevant_project_ids or row['id'] in relevant_project_ids]
    pages = [row for row in db_data['pages'] if not relevant_project_ids or row['project_id'] in relevant_project_ids]
    tasks = [row for row in db_data['tasks'] if row['id'] in allowed_tasks or row['project_id'] in relevant_project_ids]
    page_ids = {row['id'] for row in pages}
    versions = [row for row in db_data['versions'] if row['page_id'] in page_ids]

    filtered_db = {
        'projects': projects,
        'projects_by_id': {row['id']: row for row in projects},
        'pages': pages,
        'pages_by_id': {row['id']: row for row in pages},
        'pages_by_project': defaultdict(list),
        'tasks': tasks,
        'tasks_by_id': {row['id']: row for row in tasks},
        'tasks_by_project': defaultdict(list),
        'versions': versions,
        'versions_by_page': defaultdict(list),
        'current_versions': {},
    }
    for row in pages:
        filtered_db['pages_by_project'][row['project_id']].append(row)
    for row in tasks:
        filtered_db['tasks_by_project'][row['project_id']].append(row)
    for row in versions:
        filtered_db['versions_by_page'][row['page_id']].append(row)
        if row['is_current']:
            filtered_db['current_versions'][row['page_id']] = row
    for rows in filtered_db['pages_by_project'].values():
        rows.sort(key=lambda item: safe_int(item.get('order_index')) or 0)
    for rows in filtered_db['tasks_by_project'].values():
        rows.sort(key=lambda item: item.get('created_at') or '')
    for rows in filtered_db['versions_by_page'].values():
        rows.sort(key=lambda item: safe_int(item.get('version_number')) or 0)

    return filtered_entries, filtered_meta, filtered_db


def build_project_rows(db_data: dict[str, Any]) -> list[dict[str, Any]]:
    rows = []
    for project in sorted(db_data['projects'], key=lambda item: item['id']):
        project_id = project['id']
        tasks = db_data['tasks_by_project'].get(project_id, [])
        pages = db_data['pages_by_project'].get(project_id, [])
        version_count = sum(len(db_data['versions_by_page'].get(page['id'], [])) for page in pages)
        current_version_count = sum(1 for page in pages if db_data['current_versions'].get(page['id']))
        rows.append({
            'project_id': project_id,
            'project_name': project.get('project_name') or project.get('idea_prompt'),
            'creation_type': project.get('creation_type'),
            'project_status': project.get('status'),
            'restyle_prompt': project.get('restyle_prompt'),
            'total_pages': len(pages),
            'task_count': len(tasks),
            'version_count': version_count,
            'current_version_count': current_version_count,
            'has_snapshot_pages': sum(1 for page in pages if page.get('restyle_base_prompt_snapshot')),
            'first_task_created_at': tasks[0]['created_at'] if tasks else None,
            'last_task_created_at': tasks[-1]['created_at'] if tasks else None,
        })
    return rows


def build_task_rows(db_data: dict[str, Any], task_meta: dict[str, dict[str, Any]]) -> list[dict[str, Any]]:
    rows = []
    task_ids = sorted(set(task_meta) | set(db_data['tasks_by_id']))
    for task_id in task_ids:
        meta = task_meta.get(task_id, {})
        task = db_data['tasks_by_id'].get(task_id, {})
        project = db_data['projects_by_id'].get(task.get('project_id') or meta.get('project_id'), {})
        progress = {}
        if task.get('progress'):
            try:
                progress = json.loads(task['progress'])
            except json.JSONDecodeError:
                progress = {}
        rows.append({
            'task_id': task_id,
            'project_id': task.get('project_id') or meta.get('project_id'),
            'project_name': project.get('project_name') or project.get('idea_prompt'),
            'task_type': task.get('task_type'),
            'task_status': task.get('status'),
            'flow_kind': meta.get('flow_kind'),
            'request_kind': meta.get('request_kind'),
            'page_count': meta.get('page_count'),
            'completed': meta.get('completed', progress.get('completed')),
            'failed': meta.get('failed', progress.get('failed')),
            'error_message': task.get('error_message'),
            'created_at': task.get('created_at'),
            'completed_at': task.get('completed_at'),
            'debug_path': meta.get('debug_path'),
            'has_started_json': meta.get('has_started_json'),
            'has_summary_json': meta.get('has_summary_json'),
            'has_saved_version': meta.get('has_saved_version'),
        })
    return rows


def build_page_rows(db_data: dict[str, Any], uploads_root: Path) -> list[dict[str, Any]]:
    rows = []
    for page in sorted(db_data['pages'], key=lambda item: (item['project_id'], safe_int(item.get('order_index')) or 0)):
        project = db_data['projects_by_id'].get(page['project_id'], {})
        current_version = db_data['current_versions'].get(page['id'], {})
        current_image_host = host_path(current_version.get('image_path'), uploads_root)
        generated_host = host_path(page.get('generated_image_path'), uploads_root)
        cached_host = host_path(page.get('cached_image_path'), uploads_root)
        original_host = host_path(page.get('original_slide_image_path'), uploads_root)
        rows.append({
            'project_id': page['project_id'],
            'project_name': project.get('project_name') or project.get('idea_prompt'),
            'page_id': page['id'],
            'order_index': page.get('order_index'),
            'page_status': page.get('status'),
            'generated_image_path': page.get('generated_image_path'),
            'generated_image_host_path': str(generated_host) if generated_host else None,
            'generated_image_exists': generated_host.exists() if generated_host else False,
            'cached_image_path': page.get('cached_image_path'),
            'cached_image_host_path': str(cached_host) if cached_host else None,
            'cached_image_exists': cached_host.exists() if cached_host else False,
            'original_slide_image_path': page.get('original_slide_image_path'),
            'original_slide_host_path': str(original_host) if original_host else None,
            'original_slide_exists': original_host.exists() if original_host else False,
            'restyle_base_prompt_snapshot': page.get('restyle_base_prompt_snapshot'),
            'snapshot_len': len(page.get('restyle_base_prompt_snapshot') or ''),
            'current_version_number': current_version.get('version_number'),
            'current_version_image_path': current_version.get('image_path'),
            'current_version_host_path': str(current_image_host) if current_image_host else None,
            'current_version_exists': current_image_host.exists() if current_image_host else False,
            'version_count': len(db_data['versions_by_page'].get(page['id'], [])),
        })
    return rows


def match_debug_entry(entries: list[DebugEntry], page_id: str, version_number: int | None) -> DebugEntry | None:
    if version_number is not None:
        for entry in entries:
            if entry.page_id == page_id and entry.version_number == version_number:
                return entry
    matching = [entry for entry in entries if entry.page_id == page_id]
    if len(matching) == 1:
        return matching[0]
    return None


def build_version_rows(db_data: dict[str, Any], debug_entries: list[DebugEntry], uploads_root: Path) -> list[dict[str, Any]]:
    rows = []
    for version in sorted(db_data['versions'], key=lambda item: (db_data['pages_by_id'].get(item['page_id'], {}).get('project_id') or '', db_data['pages_by_id'].get(item['page_id'], {}).get('order_index') or 0, safe_int(item.get('version_number')) or 0)):
        page = db_data['pages_by_id'].get(version['page_id'], {})
        project = db_data['projects_by_id'].get(page.get('project_id'), {})
        entry = match_debug_entry(debug_entries, version['page_id'], safe_int(version.get('version_number')))
        task = db_data['tasks_by_id'].get(entry.task_id, {}) if entry else {}
        host_image = host_path(version.get('image_path'), uploads_root)
        rows.append({
            'project_id': page.get('project_id'),
            'project_name': project.get('project_name') or project.get('idea_prompt'),
            'project_status': project.get('status'),
            'page_id': page.get('id'),
            'order_index': page.get('order_index'),
            'page_status': page.get('status'),
            'version_number': version.get('version_number'),
            'is_current': bool(version.get('is_current')),
            'image_path': version.get('image_path'),
            'host_image_path': str(host_image) if host_image else None,
            'image_exists': host_image.exists() if host_image else False,
            'source_task_id': entry.task_id if entry else None,
            'source_task_type': task.get('task_type'),
            'task_status': task.get('status'),
            'request_kind': entry.request_kind if entry else None,
            'flow_kind': entry.flow_kind if entry else None,
            'source_version_number': entry.source_version_number if entry else None,
            'snapshot_persisted': entry.snapshot_persisted if entry else None,
            'snapshot_source': entry.snapshot_source if entry else None,
            'provider': entry.provider if entry else None,
            'model': entry.model if entry else None,
            'thinking_level': entry.thinking_level if entry else None,
            'elapsed_seconds': entry.elapsed_seconds if entry else None,
            'result_image_size': entry.result_image_size if entry else None,
            'error_stage': entry.error_stage if entry else None,
            'error_message': entry.error_message if entry else None,
            'prompt_full': entry.prompt_full if entry else None,
            'prompt_preview': prompt_preview(entry.prompt_full) if entry else None,
            'request_json_full': entry.request_json_full if entry else None,
            'conversation_json_full': entry.conversation_json_full if entry else None,
            'image_manifest_json_full': entry.image_manifest_json_full if entry else None,
            'turns_summary_json_full': entry.turns_summary_json_full if entry else None,
        })
    return rows


def build_request_rows(db_data: dict[str, Any], debug_entries: list[DebugEntry]) -> list[dict[str, Any]]:
    rows = []
    for entry in sorted(debug_entries, key=lambda item: (item.project_id or '', item.order_index or 0, item.version_number or 0, item.task_id)):
        project = db_data['projects_by_id'].get(entry.project_id, {})
        rows.append({
            'task_id': entry.task_id,
            'project_id': entry.project_id,
            'project_name': project.get('project_name') or project.get('idea_prompt'),
            'page_id': entry.page_id,
            'order_index': entry.order_index,
            'version_number': entry.version_number,
            'request_kind': entry.request_kind,
            'flow_kind': entry.flow_kind,
            'provider': entry.provider,
            'model': entry.model,
            'thinking_level': entry.thinking_level,
            'prompt_len': entry.prompt_len,
            'prompt_full': entry.prompt_full,
            'prompt_preview': prompt_preview(entry.prompt_full),
            'request_json_full': entry.request_json_full,
            'request_event_json_full': entry.request_event_json_full,
            'conversation_json_full': entry.conversation_json_full,
            'conversation_preview': prompt_preview(entry.prompt_full),
            'ref_image_paths_json': entry.ref_image_paths_json,
            'image_manifest_json_full': entry.image_manifest_json_full,
            'turns_summary_json_full': entry.turns_summary_json_full,
            'snapshot_source': entry.snapshot_source,
            'degraded_context': entry.degraded_context,
            'provider_fallback': entry.provider_fallback,
            'elapsed_seconds': entry.elapsed_seconds,
            'error_stage': entry.error_stage,
            'error_message': entry.error_message,
            'provider_result_json_full': entry.provider_result_json_full,
            'saved_version_json_full': entry.saved_version_json_full,
        })
    return rows


def build_flat_rows(db_data: dict[str, Any], version_rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    rows = []
    for row in version_rows:
        project = db_data['projects_by_id'].get(row['project_id'], {})
        rows.append({
            'project_id': row['project_id'],
            'project_name': row['project_name'],
            'creation_type': project.get('creation_type'),
            'project_status': row['project_status'],
            'restyle_prompt': project.get('restyle_prompt'),
            'task_id': row['source_task_id'],
            'task_type': row['source_task_type'],
            'task_status': row['task_status'],
            'flow_kind': row['flow_kind'],
            'request_kind': row['request_kind'],
            'page_id': row['page_id'],
            'order_index': row['order_index'],
            'page_status': row['page_status'],
            'version_number': row['version_number'],
            'is_current': row['is_current'],
            'source_version_number': row['source_version_number'],
            'image_path': row['image_path'],
            'host_image_path': row['host_image_path'],
            'image_exists': row['image_exists'],
            'provider': row['provider'],
            'model': row['model'],
            'thinking_level': row['thinking_level'],
            'prompt_len': len(row['prompt_full']) if row['prompt_full'] else None,
            'prompt_preview': row['prompt_preview'],
            'prompt_full': row['prompt_full'],
            'snapshot_source': row['snapshot_source'],
            'snapshot_persisted': row['snapshot_persisted'],
            'elapsed_seconds': row['elapsed_seconds'],
            'result_image_size': row['result_image_size'],
            'error_stage': row['error_stage'],
            'error_message': row['error_message'],
            'request_json_full': row['request_json_full'],
            'conversation_json_full': row['conversation_json_full'],
            'image_manifest_json_full': row['image_manifest_json_full'],
            'turns_summary_json_full': row['turns_summary_json_full'],
        })
    return rows


def write_sheet(ws, headers: list[str], rows: list[dict[str, Any]], uploads_root: Path | None = None):
    ws.append(headers)
    for cell in ws[1]:
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = WRAP_ALIGNMENT

    for row in rows:
        values = [row.get(header) for header in headers]
        ws.append(values)

    for row in ws.iter_rows(min_row=2):
        for idx, cell in enumerate(row, start=1):
            header = headers[idx - 1]
            cell.font = BODY_FONT
            cell.alignment = WRAP_ALIGNMENT if header in LONG_TEXT_COLUMNS else TOP_ALIGNMENT
            if uploads_root and header.endswith('host_path') and cell.value:
                target = Path(str(cell.value))
                if target.exists():
                    cell.hyperlink = target.resolve().as_uri()
                    cell.style = 'Hyperlink'
            if header == 'debug_path' and cell.value:
                target = Path(str(cell.value))
                if target.exists():
                    cell.hyperlink = target.resolve().as_uri()
                    cell.style = 'Hyperlink'

    ws.freeze_panes = 'A2'
    ws.auto_filter.ref = ws.dimensions
    for idx, header in enumerate(headers, start=1):
        max_len = len(str(header))
        for cell in ws[get_column_letter(idx)]:
            if cell.value is None:
                continue
            max_len = max(max_len, min(80, len(str(cell.value))))
        if header in LONG_TEXT_COLUMNS:
            width = 60
        elif header.endswith('host_path') or header.endswith('json_full'):
            width = 50
        else:
            width = min(max_len + 2, 28)
        ws.column_dimensions[get_column_letter(idx)].width = width


def save_workbook(path: Path, sheets: list[tuple[str, list[str], list[dict[str, Any]]]], uploads_root: Path | None = None):
    wb = Workbook()
    default_ws = wb.active
    wb.remove(default_ws)
    for name, headers, rows in sheets:
        ws = wb.create_sheet(title=name)
        write_sheet(ws, headers, rows, uploads_root=uploads_root)
    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)


def export_debug_workbooks(
    *,
    debug_root: Path,
    db_path: Path,
    uploads_root: Path,
    output_dir: Path,
    project_ids: list[str] | None = None,
    task_ids: list[str] | None = None,
    latest: int | None = None,
) -> dict[str, str]:
    db_data = load_database(db_path)
    debug_entries, task_meta = load_debug_entries(debug_root)
    filtered_entries, filtered_task_meta, filtered_db = apply_filters(
        db_data,
        debug_entries,
        task_meta,
        set(project_ids or []) or None,
        set(task_ids or []) or None,
        latest,
    )

    project_rows = build_project_rows(filtered_db)
    task_rows = build_task_rows(filtered_db, filtered_task_meta)
    page_rows = build_page_rows(filtered_db, uploads_root)
    version_rows = build_version_rows(filtered_db, filtered_entries, uploads_root)
    request_rows = build_request_rows(filtered_db, filtered_entries)
    flat_rows = build_flat_rows(filtered_db, version_rows)

    overview_path = output_dir / 'restyle-debug-overview.xlsx'
    flat_path = output_dir / 'restyle-debug-flat.xlsx'
    save_workbook(
        overview_path,
        [
            ('Projects', PROJECT_HEADERS, project_rows),
            ('Tasks', TASK_HEADERS, task_rows),
            ('Pages', PAGE_HEADERS, page_rows),
            ('Versions', VERSION_HEADERS, version_rows),
            ('Requests', REQUEST_HEADERS, request_rows),
        ],
        uploads_root=uploads_root,
    )
    save_workbook(flat_path, [('FlatExport', FLAT_HEADERS, flat_rows)], uploads_root=uploads_root)

    return {'overview': str(overview_path), 'flat': str(flat_path)}


def main() -> int:
    args = parse_args()
    outputs = export_debug_workbooks(
        debug_root=Path(args.debug_root).expanduser().resolve(),
        db_path=Path(args.db).expanduser().resolve(),
        uploads_root=Path(args.uploads_root).expanduser().resolve(),
        output_dir=Path(args.output_dir).expanduser().resolve(),
        project_ids=args.project_ids,
        task_ids=args.task_ids,
        latest=args.latest,
    )
    print(json.dumps(outputs, ensure_ascii=False, indent=2))
    return 0


if __name__ == '__main__':
    raise SystemExit(main())
